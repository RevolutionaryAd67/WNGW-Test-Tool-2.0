"""Utility helpers to parse Excel signal lists."""
from __future__ import annotations

from pathlib import Path
from typing import List

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime
    load_workbook = None

REQUIRED_COLUMNS = ["Datenpunkt", "IOA 3", "IOA 2", "IOA 1"]


def parse_signal_excel(path: Path) -> List[dict]:
    """Parse the Excel file and return a list of signal definitions."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse Excel files")

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [header for header in REQUIRED_COLUMNS if header not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    column_index = {header: headers.index(header) for header in headers}
    signals: List[dict] = []
    for row in sheet.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue
        signal = {
            "name": row[column_index["Datenpunkt"]].value,
            "ioa": [
                int(row[column_index["IOA 3"]].value or 0),
                int(row[column_index["IOA 2"]].value or 0),
                int(row[column_index["IOA 1"]].value or 0),
            ],
            "raw": {headers[idx]: cell.value for idx, cell in enumerate(row)},
        }
        signals.append(signal)
    return signals
